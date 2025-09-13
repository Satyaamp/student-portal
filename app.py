from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE = 'students.xlsx'

# ----------------- Get Student -----------------
def get_student(enroll_no):
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    enroll_no = str(enroll_no).strip()  # ensure string
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if str(row[0].value).strip() == enroll_no:
            return [cell.value for cell in row]
    return None

# ----------------- Update Student -----------------
def update_student(enroll_no, updated_data):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        enroll_no = str(enroll_no).strip()
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value).strip() == enroll_no:
                for i in range(len(updated_data)):
                    row[i].value = updated_data[i]
                wb.save(EXCEL_FILE)
                return True
    except PermissionError:
        print("ERROR: Cannot save the Excel file. Make sure it is closed in Excel.")
        return False
    return False

# ----------------- Home / Search -----------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        enroll_no = request.form.get('enroll_no')
        if not enroll_no:
            return "Enrollment number is missing", 400

        student = get_student(enroll_no)
        if student:
            headers = ['Enrollment', 'Name', 'DOB', 'Address', 'Contact', 'Course', 'Year', 'Email', 'Fee Status', 'Remarks']
            cleaned_headers = [h.lower().replace(' ', '_') for h in headers]
            return render_template('student_info.html', student=student, headers=headers, cleaned_headers=cleaned_headers)
        else:
            return "Enrollment number not found"
    return render_template('index.html')

# ----------------- Update Student -----------------
@app.route('/update', methods=['POST'])
def update():
    updated_data = [
        request.form['enrollment'],
        request.form['name'],
        request.form['dob'],
        request.form['address'],
        request.form['contact'],
        request.form['course'],
        request.form['year'],
        request.form['email'],
        request.form['fee_status'],
        request.form['remarks']
    ]
    success = update_student(updated_data[0], updated_data)
    if success:
        return render_template('confirmation.html', student=updated_data)
    else:
        return "Error updating student. Ensure Excel file is closed.", 500

# ----------------- Run Flask -----------------
if __name__ == '__main__':
    app.run(debug=True)
